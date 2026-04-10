"""
export_signals.py — Export active signal data to Excel.

For each unique market (most recent signal row per market_id), exports:
  question, category, market_price, claude_prob, gap, days_to_resolution,
  confidence, resolved_value, was_claude_correct, timestamp

Filters to markets in the 10-90% active signal band.
Sorted by gap (claude_prob - market_price) descending.

Output: signals_export.xlsx in the project folder.
"""

import os
import sqlite3
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

DB_PATH  = os.getenv("DB_PATH", "signals.db")
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "signals_export.xlsx")

MIN_PRICE = 0.10
MAX_PRICE = 0.90

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
COL_HEADER_BG  = "1F4E79"   # dark navy
COL_HEADER_FG  = "FFFFFF"
COL_POSITIVE   = "C6EFCE"   # green tint  — claude higher than market
COL_NEGATIVE   = "FFDDC1"   # orange tint — claude lower than market
COL_CORRECT    = "C6EFCE"
COL_WRONG      = "FFC7CE"
COL_NEUTRAL    = "FFFFFF"
COL_ALT_ROW    = "F2F7FC"   # light blue alternating row


def _pct(v):
    """Format float as percentage string for display."""
    if v is None:
        return None
    return round(float(v) * 100, 1)


def fetch_data(conn: sqlite3.Connection) -> list[dict]:
    conn.row_factory = sqlite3.Row
    # For unresolved markets: filter by current (most recent) market_price.
    # For resolved markets:   filter by entry (first) market_price — the post-
    #                         resolution price is always near 0 or 1 and would
    #                         otherwise exclude every resolved row.
    rows = conn.execute("""
        SELECT
            latest.market_id,
            latest.question,
            latest.category,
            latest.market_price,
            latest.claude_prob,
            latest.days_to_resolution,
            latest.confidence,
            latest.resolved_value,
            latest.was_claude_correct,
            latest.timestamp
        FROM signals latest
        -- Most recent row per market
        WHERE latest.id IN (SELECT MAX(id) FROM signals GROUP BY market_id)
          AND latest.claude_prob IS NOT NULL
          AND (
              -- Unresolved: current price must be in the active band
              (latest.resolved_value IS NULL
               AND latest.market_price BETWEEN ? AND ?)
              OR
              -- Resolved: entry price (first signal) must have been in the band
              (latest.resolved_value IS NOT NULL
               AND (SELECT s2.market_price
                    FROM signals s2
                    WHERE s2.market_id = latest.market_id
                    ORDER BY s2.id ASC LIMIT 1) BETWEEN ? AND ?)
          )
        ORDER BY (latest.claude_prob - latest.market_price) DESC
    """, (MIN_PRICE, MAX_PRICE, MIN_PRICE, MAX_PRICE)).fetchall()
    return [dict(r) for r in rows]


def build_workbook(rows: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signals"

    # ── Metadata row ────────────────────────────────────────────────────────
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A1"] = f"Polymarket Signal Export  |  {len(rows)} markets  |  Generated {now_str}"
    ws["A1"].font = Font(bold=True, size=11)
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 18

    # ── Column definitions ───────────────────────────────────────────────────
    columns = [
        ("Question",          60),
        ("Category",          18),
        ("Mkt Price",         10),
        ("Claude Prob",       11),
        ("Gap",               10),
        ("Days to Res.",      12),
        ("Confidence",        11),
        ("Resolved",           9),
        ("Claude Correct",    14),
        ("Timestamp (UTC)",   20),
    ]

    header_fill   = PatternFill("solid", fgColor=COL_HEADER_BG)
    header_font   = Font(bold=True, color=COL_HEADER_FG, size=10)
    thin_side     = Side(style="thin", color="CCCCCC")
    thin_border   = Border(bottom=thin_side)

    # ── Header row ───────────────────────────────────────────────────────────
    HEADER_ROW = 2
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 20

    # Freeze header + metadata
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # ── Data rows ─────────────────────────────────────────────────────────────
    pct_fmt   = "0.0%"
    date_fmt  = "YYYY-MM-DD HH:MM"

    for row_idx, r in enumerate(rows, start=HEADER_ROW + 1):
        alt = (row_idx % 2 == 0)
        base_fill = PatternFill("solid", fgColor=COL_ALT_ROW) if alt else None

        mp  = r["market_price"]
        cp  = r["claude_prob"]
        gap = (cp - mp) if (cp is not None and mp is not None) else None

        # Resolved / correct labels
        if r["resolved_value"] is None:
            resolved_label = "Open"
        elif r["resolved_value"] == 1.0:
            resolved_label = "YES"
        else:
            resolved_label = "NO"

        correct_raw = r["was_claude_correct"]
        if correct_raw == 1:
            correct_label = "Correct"
        elif correct_raw == 0:
            correct_label = "Wrong"
        else:
            correct_label = ""

        # Timestamp — strip timezone suffix for Excel
        ts_raw = (r["timestamp"] or "")[:19].replace("T", " ")

        values = [
            r["question"],
            r["category"],
            mp,
            cp,
            gap,
            round(r["days_to_resolution"], 1) if r["days_to_resolution"] is not None else None,
            (r["confidence"] or "").capitalize(),
            resolved_label,
            correct_label,
            ts_raw,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 1))
            cell.border = Border(bottom=Side(style="thin", color="EEEEEE"))

            # Apply base alternating fill unless overridden below
            if base_fill:
                cell.fill = base_fill

        # Colour-code Gap column (col 5)
        gap_cell = ws.cell(row=row_idx, column=5)
        if gap is not None:
            gap_cell.number_format = "+0.0%;-0.0%;0.0%"
            if gap > 0.05:
                gap_cell.fill = PatternFill("solid", fgColor=COL_POSITIVE)
            elif gap < -0.05:
                gap_cell.fill = PatternFill("solid", fgColor=COL_NEGATIVE)

        # Percentage format for price columns
        for col_idx in (3, 4):
            ws.cell(row=row_idx, column=col_idx).number_format = pct_fmt

        # Colour-code Correct column (col 9)
        correct_cell = ws.cell(row=row_idx, column=9)
        if correct_raw == 1:
            correct_cell.fill = PatternFill("solid", fgColor=COL_CORRECT)
            correct_cell.font = Font(bold=True)
        elif correct_raw == 0:
            correct_cell.fill = PatternFill("solid", fgColor=COL_WRONG)
            correct_cell.font = Font(bold=True)

        ws.row_dimensions[row_idx].height = 30 if len(r["question"]) > 80 else 18

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(columns))}{HEADER_ROW + len(rows)}"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16

    def s_header(row, label):
        c = ws2.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, color=COL_HEADER_FG)
        c.fill = PatternFill("solid", fgColor=COL_HEADER_BG)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def s_row(row, label, value):
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=value).alignment = Alignment(horizontal="right")

    in_band   = [r for r in rows if r["claude_prob"] is not None]
    resolved  = [r for r in in_band if r["resolved_value"] is not None]
    correct   = [r for r in resolved if r["was_claude_correct"] == 1]
    bullish   = [r for r in in_band if r["claude_prob"] is not None and r["market_price"] is not None and r["claude_prob"] - r["market_price"] > 0.05]
    bearish   = [r for r in in_band if r["claude_prob"] is not None and r["market_price"] is not None and r["market_price"] - r["claude_prob"] > 0.05]

    s_header(1, "Overview")
    s_row(2,  "Export generated",        now_str)
    s_row(3,  "Markets in export",       len(rows))
    s_row(4,  "Resolved markets",        len(resolved))
    s_row(5,  "Claude accuracy",         f"{len(correct)/len(resolved)*100:.1f}%" if resolved else "N/A")
    s_header(7, "Signal Direction")
    s_row(8,  "Claude > Market (+5%)",   len(bullish))
    s_row(9,  "Claude < Market (-5%)",   len(bearish))
    s_row(10, "No clear edge",           len(in_band) - len(bullish) - len(bearish))
    s_header(12, "Category Breakdown")

    from collections import Counter
    cats = Counter(r["category"] for r in rows)
    row_n = 13
    for cat, n in sorted(cats.items(), key=lambda x: -x[1]):
        s_row(row_n, cat, n)
        row_n += 1

    return wb


def main():
    if not os.path.exists(DB_PATH):
        print(f"ERROR: DB not found at {DB_PATH}")
        raise SystemExit(1)

    conn = sqlite3.connect(DB_PATH)
    rows = fetch_data(conn)
    conn.close()

    print(f"Fetched {len(rows)} markets from {DB_PATH}")

    wb = build_workbook(rows)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
